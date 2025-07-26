import json
import os
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
import locale
from datetime import datetime
from pdf_processor import PdfProcessor
# locale.setlocale(locale.LC_TIME, 'bg_BG.UTF-8')  # Set to Bulgarian locale
bg_months = {
    'Януари': 'January',
    'Февруари': 'February',
    'Март': 'March',
    'Април': 'April',
    'Май': 'May',
    'Юни': 'June',
    'Юли': 'July',
    'Август': 'August',
    'Септември': 'September',
    'Октомври': 'October',
    'Ноември': 'November',
    'Декември': 'December',
    'януари': 'january',
    'февруари': 'february',
    'март': 'march',
    'април': 'april',
    'май': 'may',
    'юни': 'june',
    'юли': 'july',
    'август': 'august',
    'септември': 'september',
    'октомври': 'october',
    'ноември': 'november',
    'декември': 'december'
}

def bg_to_en_month(date_str):
    for bg, en in bg_months.items():
        if bg in date_str:
            return date_str.replace(bg, en)
    return date_str

def process_pdfs(directory):
    """Обработва всички PDF файлове и съхранява данните в речник."""
    data_by_object_code = {}

    for pdf_file in os.listdir(directory):
        if pdf_file.endswith(".pdf"):
            pdf_path = os.path.join(directory, pdf_file)

            # Симулиране на извличане на текст от PDF (заменете с реална логика)
            #extracted_text = simulate_pdf_extraction(pdf_path)
            pdf_processor = PdfProcessor()
            extracted_text = pdf_processor.extract_text(pdf_path)      
            extracted_text = extracted_text.encode('latin1').decode('windows-1251')
            print("############################################")
            print("Обработване на PDF файл:", pdf_file)
            # Запис във файл
            txt_filename = pdf_file + ".txt"
            with open(txt_filename, "w", encoding="utf-8") as f:
                f.write(extracted_text)
            if extracted_text:
                current_object_name = None
                current_object_address = None
                current_month = None
                current_energy_sum = 0.0
                current_energy_sum2 = 0.0
                current_energy_sum3 = 0.0
                find_date_doc = "Основание: Електрическа енергия за месец"
                find_str_con1 = "Достъп високо напрежение"
                find_str_con2 = "Общо сума"
                find_str_con3 = "Надбавка за използвана реактивна енергия"
                match_sum = False
                blocks = [[]]
                blocks.append([])  # Начало на нов блок
                blockindex = 0
                for line in extracted_text.splitlines():
                    line = line.strip()
                    blocks[blockindex].append(line)
                    if  line.startswith("- - - - "):
                        blockindex += 1
                        blocks.append([])

                    if line.startswith("Наименование на обекта:"):
                        current_object_name = line.replace("Наименование на обекта:", "").strip()
                        if current_object_name:
                            parts = current_object_name.rsplit(" ", 1)
                            if len(parts) == 2:
                                object_code = parts[1]
                                object_name = parts[0]
                            else:
                                object_code = current_object_name
                                object_name = ""
                    elif line.startswith("Адрес на обекта:"):
                        current_object_address = line.replace("Адрес на обекта: ", "").strip()
                        current_object_address = current_object_address.replace("Кодов номер:", "").strip()
                    elif line.startswith("Основание: Електрическа енергия за месец"):
                        current_month = line.replace("Основание: Електрическа енергия за месец", "").strip()
                        print("Обработване на месец:", current_month)
                        current_month = bg_to_en_month(current_month)

                        current_month = datetime.strptime(current_month, "%B %Y")  # Преобразуване в дата
                    elif line.startswith(find_str_con1):
                        line = line.replace(find_str_con1, "").strip()
                        if "кВтч" in line:
                            energy_part = line.split("кВтч")[0].strip()
                            energy_part = energy_part.replace(" ", "")
                            energy_part = energy_part[::-1]
                            energy_part = energy_part.lstrip("0")
                            try:
                                current_energy_sum += float(energy_part)
                                
                            except ValueError:
                                pass
                    elif line.startswith(find_str_con2):
                        line = line.replace(find_str_con2, "").strip()
                        line = line.replace(",", "").strip()
                        energy_part = line
                        energy_part = energy_part.replace(" ", "")
                        energy_part = energy_part[::-1]
                        energy_part = energy_part.lstrip("0")
                        try:
                            current_energy_sum2 += float(energy_part)
                            match_sum = True
                        except ValueError:
                            pass
                    elif line.startswith(find_str_con3):
                        line = line.replace(find_str_con3, "").strip()
                        if "кВАрч" in line:
                            energy_part = line.split("кВАрч")[0].strip()
                            energy_part = energy_part.replace(" ", "")
                            energy_part = energy_part[::-1]
                            energy_part = energy_part.lstrip("0")
                        try:
                            current_energy_sum3 += float(energy_part)
                        except ValueError:
                            pass
                    if  match_sum and line.startswith("- - - - "):
                        if object_code not in data_by_object_code:
                            data_by_object_code[object_code] = {
                                "object_name": current_object_name,
                                "object_address": current_object_address,
                                "rows": []
                            }
                        data_by_object_code[object_code]["rows"].append([
                            pdf_file, current_month, current_energy_sum, current_energy_sum3
                        ])
                        current_energy_sum = 0.0
                        current_energy_sum3 = 0.0
                        current_energy_sum2 = 0.0
                        match_sum = False
                # Записване на блоковете във файл
                blocks.remove(blocks[-1])  # Премахване на последния празен блок
                with open("arr"+txt_filename, "w", encoding="utf-8") as f:
                    for block in blocks:
                        f.write("\n-----block "+str(blocks.index(block))+"-----\n"  )
                        if block:
                            f.write(block.join("\n") )
    return data_by_object_code

def generate_excel(data_by_object_code, excel_path):
    """Генерира Excel файл с таблици и графики за всеки обект."""
    workbook = Workbook()

    for object_code, data in data_by_object_code.items():
        sheet = workbook.create_sheet(title=object_code[:31])  # Ограничение на имената на sheet-овете до 31 символа

        # Добавяне на антетка
        sheet.append(["Код на обекта:", object_code])
        sheet.append(["Име на обекта:", data["object_name"]])
        sheet.append(["Адрес на обекта:", data["object_address"]])
        sheet.append([])  # Празен ред за разделяне

        # Добавяне на заглавия на таблицата
        sheet.append(["PDF Path", "За месец (Дата)", "Активна мощност (W)", "Реактивна мощност (W)"])

        # Добавяне на редовете
        for row in data["rows"]:
            sheet.append([row[0], row[1].strftime("%Y-%m"), row[2], row[3]])

        # # Създаване на графика
        # chart = LineChart()
        # chart.title = "Активна и Реактивна мощност по месеци"
        # chart.style = 13
        # chart.x_axis.title = "Месец"
        # chart.y_axis.title = "Мощност (W)"

        # # Данни за графиката
        # data = Reference(sheet, min_col=3, max_col=4, min_row=6, max_row=sheet.max_row)
        # categories = Reference(sheet, min_col=2, min_row=6, max_row=sheet.max_row)  # Колона B за "Месец"
        # chart.add_data(data, titles_from_data=True)
        # chart.set_categories(categories)

        # # Поставяне на графиката под таблицата
        # sheet.add_chart(chart, f"A{sheet.max_row + 2}")

    # Премахване на празния sheet, ако съществува
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    workbook.save(excel_path)

def simulate_pdf_extraction(pdf_path):
    """Симулира извличане на текст от PDF (заменете с реална логика)."""
    return f"""
    Наименование на обекта: Обект 1
    Адрес на обекта: ул. Примерна 1
    Основание: Електрическа енергия за месец Януари 2025
    100 кВтч
    50 кВАрч
    """

def main():
    parser = argparse.ArgumentParser(description="Process multiple PDF files and save data to an Excel file.")
    parser.add_argument("excel_path", nargs="?", default="test.xlsx", help="Path to the output Excel file.")
    parser.add_argument("pdf_directory", nargs="?", default="test", help="Path to the directory containing PDF files.")
    args = parser.parse_args()

    if not os.path.isdir(args.pdf_directory):
        print("The specified directory does not exist.")
        return

    # Обработка на PDF файловете
    data_by_object_code = process_pdfs(args.pdf_directory)

    # Генериране на Excel файла
    generate_excel(data_by_object_code, args.excel_path)
    # След като имаш data_by_object_code:
    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(data_by_object_code, f, ensure_ascii=False, default=str, indent=2)
if __name__ == "__main__":
    main()