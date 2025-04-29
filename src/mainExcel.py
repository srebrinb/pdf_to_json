import os
import argparse
from openpyxl import Workbook, load_workbook
from pdf_processor import PdfProcessor

def parse_text_to_excel(extracted_text, excel_path, pdf_path):
    rows_by_object_code = {}
    current_object_name_old = None
    current_object_name = None
    current_object_address = None
    current_month = None
    current_energy_sum = 0.0
    current_energy_sum2 = 0.0
    current_energy_sum3 = 0.0
    find_date_doc = "Основание: Електрическа енергия за месец"
    find_str_con1 = "Достъп високо напрежение"
    find_str_con2 = "Общо сума"
    find_str_con3 = "Надбавка за използвана реактивна енергия"
    match_sum = False

    for line in extracted_text.splitlines():
        line = line.strip()
        if line.startswith("Наименование на обекта:"):
            current_object_name = line.replace("Наименование на обекта:", "").strip()
            if current_object_name:
                parts = current_object_name.rsplit(" ", 1)
                if len(parts) == 2:
                    object_code = parts[1]
                    object_name = parts[0]
                else:
                    object_code = current_object_name
                    object_name = ""
        elif line.startswith("Адрес на обекта: "):
            current_object_address = line.replace("Адрес на обекта: ", "").strip()
            current_object_address = current_object_address.replace("Кодов номер:", "").strip()
        elif line.startswith(find_date_doc):
            current_month = line.replace(find_date_doc, "").strip()

        if line.startswith(find_str_con1):
            line = line.replace(find_str_con1, "").strip()
            if "кВтч" in line:
                energy_part = line.split("кВтч")[0].strip()
                energy_part = energy_part.replace(" ", "")
                energy_part = energy_part[::-1]
                energy_part = energy_part.lstrip("0")
                try:
                    current_energy_sum += float(energy_part)
                    match_sum = True
                except ValueError:
                    pass
        if line.startswith(find_str_con2):
            line = line.replace(find_str_con2, "").strip()
            line = line.replace(",", "").strip()
            energy_part = line
            energy_part = energy_part.replace(" ", "")
            energy_part = energy_part[::-1]
            energy_part = energy_part.lstrip("0")
            try:
                current_energy_sum2 += float(energy_part)
            except ValueError:
                pass
        if line.startswith(find_str_con3):
            line = line.replace(find_str_con3, "").strip()
            if "кВАрч" in line:
                energy_part = line.split("кВАрч")[0].strip()
                energy_part = energy_part.replace(" ", "")
                energy_part = energy_part[::-1]
                energy_part = energy_part.lstrip("0")
            try:
                current_energy_sum3 += float(energy_part)
            except ValueError:
                pass

        if current_energy_sum2:
            if object_code not in rows_by_object_code:
                rows_by_object_code[object_code] = []
            rows_by_object_code[object_code].append([
                pdf_path, object_code, object_name, current_object_address,
                current_month, current_energy_sum, current_energy_sum3, current_energy_sum2
            ])
            current_energy_sum = 0.0
            current_energy_sum3 = 0.0
            current_energy_sum2 = 0.0
            match_sum = False

    # Зареждане на съществуващия Excel файл или създаване на нов
    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path)
    else:
        workbook = Workbook()

    for object_code, rows in rows_by_object_code.items():
        if object_code in workbook.sheetnames:
            sheet = workbook[object_code]
        else:
            sheet = workbook.create_sheet(title=object_code[:31])  # Ограничение на имената на sheet-овете до 31 символа
            sheet.append(["From File", "Код на обекта", "Име на обекта", "Адрес на обекта", "За месец", "активна мощност", "реактивна мощност", find_str_con2, "обща мощност","коефициент на мощността (cosφ)","CO₂ емисии (kg)"])

        # Добавяне на новите редове към съществуващия sheet
        for row in rows:
            sheet.append(row)

        # Добавяне на формулата в последната колона
        for row_idx in range(2, sheet.max_row + 1):  # Пропускаме заглавния ред
            energy_sum_cell = f"F{row_idx}"  # Колона F за current_energy_sum
            energy_sum3_cell = f"G{row_idx}"  # Колона G за current_energy_sum3
            formula_cell = f"I{row_idx}"  # Колона I за формулата
            sheet[formula_cell] = f"=SQRT({energy_sum_cell}^2 + {energy_sum3_cell}^2)"
            sheet[f"J{row_idx}"] = f"={energy_sum_cell} / SQRT({energy_sum_cell}^2 +{energy_sum3_cell}^2 )"
            sheet[f"K{row_idx}"] = f"= SQRT({energy_sum_cell}^2 + {energy_sum3_cell}^2) * 0.3"

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]  # Премахване на празния sheet, ако съществува

    workbook.save(excel_path)

def main():
    # Добавяне на argparse за подаване на параметри от командния ред
    parser = argparse.ArgumentParser(description="Extract data from a PDF and save it to an Excel file.")
    parser.add_argument("pdf_path", nargs="?", default="C:\\work\\PDF_Extr\\1.pdf", help="Path to the input PDF file (default: C:\\work\\PDF_Extr\\1.pdf)")
    parser.add_argument("excel_path", nargs="?", default="C:\\work\\PDF_Extr\\1.xlsx", help="Path to the output Excel file (default: C:\\work\\PDF_Extr\\1.xlsx)")

    try:
        args = parser.parse_args()
        pdf_path = args.pdf_path
        excel_path = args.excel_path
    except Exception as e:
        print(f"Error: Unable to parse arguments. Details: {e}")
        return

    if not os.path.exists(pdf_path):
        print("The specified PDF file does not exist.")
        return

    pdf_processor = PdfProcessor()
    extracted_text = pdf_processor.extract_text(pdf_path)

    # Decode the text from windows-1251 and re-encode to UTF-8
    try:
        extracted_text = extracted_text.encode('latin1').decode('windows-1251')
    except UnicodeError as e:
        print(f"Error: Unable to decode text. Details: {e}")
        return

    if extracted_text:
        # Парсиране на текста и запис в Excel
        parse_text_to_excel(extracted_text, excel_path, pdf_path)
        print(f"Data has been written to {excel_path}")
    else:
        print("No text extracted from the PDF.")

if __name__ == "__main__":
    main()