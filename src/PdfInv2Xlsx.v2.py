import json
import os
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference

import locale
from datetime import datetime
from pdf_processor import PdfProcessor
import zlib
# locale.setlocale(locale.LC_TIME, 'bg_BG.UTF-8')  # Set to Bulgarian locale
bg_months = {
    'Януари': 'January',
    'Февруари': 'February',
    'Март': 'March',
    'Април': 'April',
    'Май': 'May',
    'Юни': 'June',
    'Юли': 'July',
    'Август': 'August',
    'Септември': 'September',
    'Октомври': 'October',
    'Ноември': 'November',
    'Декември': 'December',
    'януари': 'january',
    'февруари': 'february',
    'март': 'march',
    'април': 'april',
    'май': 'may',
    'юни': 'june',
    'юли': 'july',
    'август': 'august',
    'септември': 'september',
    'октомври': 'october',
    'ноември': 'november',
    'декември': 'december'
}

def bg_to_en_month(date_str):
    for bg, en in bg_months.items():
        if bg in date_str:
            return date_str.replace(bg, en)
    return date_str

def process_pdfs(directory):
    """Обработва всички PDF файлове и съхранява данните в речник."""
    data_by_object_code = {}

    for pdf_file in os.listdir(directory):
        if pdf_file.endswith(".pdf"):
            pdf_path = os.path.join(directory, pdf_file)

            # Симулиране на извличане на текст от PDF (заменете с реална логика)
            #extracted_text = simulate_pdf_extraction(pdf_path)
            pdf_processor = PdfProcessor()
            extracted_text = pdf_processor.extract_text(pdf_path)      
            extracted_text = extracted_text.encode('latin1').decode('windows-1251')
            print("############################################")
            print("Обработване на PDF файл:", pdf_file)
            # Запис във файл
            # txt_filename = pdf_file + ".txt"
            # with open(txt_filename, "w", encoding="utf-8") as f:
            #     f.write(extracted_text)
            if extracted_text:
                current_object_name = None
                current_object_address = None
                current_month = None
                current_month_bg = None
                current_energy_sum = 0.0
                current_energy_sum2 = 0.0
                current_energy_sum3 = 0.0
                find_date_doc = "Основание: Електрическа енергия за месец"
                find_str_con1 = "Достъп високо напрежение"
                find_str_con2 = "Общо сума"
                find_str_con3 = "Надбавка за използвана реактивна енергия"
                match_sum = False
                blocks = extracted_text.split("- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -")
                for block in blocks:
                    for line in block.splitlines():
                        line = line.strip()
                        if line.startswith("Наименование на обекта:"):
                            current_object_name = line.replace("Наименование на обекта:", "").strip()
                            if current_object_name:
                                parts = current_object_name.rsplit(" ", 1)
                                if len(parts) == 2:
                                    object_code = parts[1]
                                    object_name = parts[0]
                                else:
                                    object_code = current_object_name
                                    object_name = ""
                        elif line.startswith("Адрес на обекта:"):
                            current_object_address = line.replace("Адрес на обекта: ", "").strip()
                            current_object_address = current_object_address.replace("Кодов номер:", "").strip()
                        elif line.startswith("Основание: Електрическа енергия за месец"):
                            current_month = line.replace("Основание: Електрическа енергия за месец", "").strip()
                            print("Обработване на месец:", current_month)
                            current_month_bg = current_month
                            current_month = bg_to_en_month(current_month)

                            current_month = datetime.strptime(current_month, "%B %Y")  # Преобразуване в дата
                        elif line.startswith(find_str_con1):
                            line = line.replace(find_str_con1, "").strip()
                            if "кВтч" in line:
                                energy_part = line.split("кВтч")[0].strip()
                                energy_part = energy_part.replace(" ", "")
                                energy_part = energy_part[::-1]
                                energy_part = energy_part.lstrip("0")
                                try:
                                    current_energy_sum += float(energy_part)
                                    
                                except ValueError:
                                    pass
                        elif line.startswith(find_str_con2):
                            line = line.replace(find_str_con2, "").strip()
                            line = line.replace(",", "").strip()
                            energy_part = line
                            energy_part = energy_part.replace(" ", "")
                            energy_part = energy_part[::-1]
                            energy_part = energy_part.lstrip("0")
                            try:
                                current_energy_sum2 += float(energy_part)
                                match_sum = True
                            except ValueError:
                                pass
                        elif line.startswith(find_str_con3):
                            line = line.replace(find_str_con3, "").strip()
                            if "кВАрч" in line:
                                energy_part = line.split("кВАрч")[0].strip()
                                energy_part = energy_part.replace(" ", "")
                                energy_part = energy_part[::-1]
                                energy_part = energy_part.lstrip("0")
                            try:
                                current_energy_sum3 += float(energy_part)
                            except ValueError:
                                pass
                    if  match_sum:
                        if object_code not in data_by_object_code:
                            data_by_object_code[object_code] = {
                                "object_code": object_code,
                                "object_name": object_name,
                                "object_address": current_object_address,
                                "rows": []
                            }
                        data_by_object_code[object_code]["rows"].append([
                            pdf_file, current_month, current_energy_sum, current_energy_sum3,current_energy_sum2,current_month_bg
                        ])
                        current_energy_sum = 0.0
                        current_energy_sum3 = 0.0
                        current_energy_sum2 = 0.0
                        match_sum = False

    return data_by_object_code

def generate_excel(data_by_object_code, excel_path):
    now = datetime.now()
    """Генерира Excel файл с таблици и графики за всеки обект."""
    workbook = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
    
    if "objects" in workbook.sheetnames:
        sheetObj = workbook["objects"]
    else:
        sheetObj = workbook.create_sheet("objects")
        sheetObj.append(["Код на обекта", "Име на обекта", "Адрес на обекта", "Брой персонал"])

    for object_code, data in data_by_object_code.items():
        if object_code in workbook.sheetnames:
            sheet = workbook[object_code]
        else:
            sheet = workbook.create_sheet(title=object_code[:31])  # Ограничение на имената на sheet-овете до 31 символа
            # Добавяне на антетка
            sheet.append(["Код на обекта:", object_code])
            sheet.append(["Име на обекта:", data["object_name"]])
            sheet.append(["Адрес на обекта:", data["object_address"]])
            sheet.append([])  # Празен ред за разделяне       
                    # Добавяне на заглавия на таблицата
            sheet.append(["PDF Path", "За месец (Дата)", "Активна мощност (W)", "Реактивна мощност (W)","Обща сума",
                           "Обща мощност (формула)","Коефициент на мощността (cosφ)", "CO₂ емисии (kg)","брой персонал", "Average emissions per employee", "Напрежение", "Ток", "Фазов ъгъл"
                          ])
            # Добави заглавията в AA и AB
            sheet.cell(row=5, column=27).value = "checksum"
            sheet.cell(row=5, column=28).value = "Дата на запис"


        
   
        # existing_keys = set(row_obj[0].value for row_obj in sheetObj.iter_rows(min_row=2))
        # if object_code not in existing_keys:
        #     # Add hyperlink with style
        #     row_idx = sheetObj.max_row + 1
        #     cell = sheetObj.cell(row=row_idx, column=1)
        #     cell.value = f'=HYPERLINK("#\'{object_code[:31]}\'!A1", "{object_code}")'
        #     cell.font = Font(color="0000FF", underline="single")
        #     sheetObj.cell(row=row_idx, column=2).value = data["object_name"]
        #     sheetObj.cell(row=row_idx, column=3).value = data["object_address"]
        #     sheetObj.cell(row=row_idx, column=4).value = sheet["I6"].value

        staff_ = -1
        for row in sheetObj.iter_rows(min_row=2):  # Прескачаме заглавията
            if object_code in str(row[0].value):  # колона 1 (A)
                staff_ = row[3].value
                break
        if staff_ == -1:
            print(f"Не е намерен персонал за обект {object_code}.")
            # Add hyperlink with style
            row_idx = sheetObj.max_row + 1
            cell = sheetObj.cell(row=row_idx, column=1)
            cell.value = f'=HYPERLINK("#\'{object_code[:31]}\'!A1", "{object_code}")'
            cell.font = Font(color="FF0000", underline="single")
            sheetObj.cell(row=row_idx, column=2).value = data["object_name"]
            sheetObj.cell(row=row_idx, column=3).value = data["object_address"]    

        # Добавяне на редовете
        for row in data["rows"]:
             # Collect existing hash values from the per-object sheet (skip header rows)
            existing_hash_keys = set()
            for r in sheet.iter_rows(min_row=5):  # Data starts from row 6
                if len(r) > 27 and r[27].value is not None:
                    existing_hash_keys.add(r[27].value)
            hash_value = zlib.crc32(str(row[1].strftime("%Y-%m")).encode("utf-8")+
                                    str(row[2]).encode("utf-8")+
                                    str(row[3]).encode("utf-8"))
           

            if hash_value not in existing_hash_keys:
                row_data = [row[0], row[5], row[2], row[3], row[4]]

                sheet.append(row_data)
                # Добави checksum и дата на запис в AA и AB
                row_idx = sheet.max_row
                sheet.cell(row=row_idx, column=27).value = hash_value
                sheet.cell(row=row_idx, column=28).value = now.strftime("%Y-%m-%d %H:%M:%S")
                for col in range(6, sheetObj.max_column + 1):
                    source_cell = sheetObj.cell(row=2, column=col)
                    formula = source_cell.value
                    target_cell = sheet.cell(row=row_idx, column=col)
                    if isinstance(formula, str):
                        formula = formula.replace("C2", f"C{row_idx}")
                        formula = formula.replace("D2", f"D{row_idx}")
                        formula = formula.replace("K2", f"K{row_idx}")
                        formula = formula.replace("F2", f"F{row_idx}")
                        formula = formula.replace("I2", f"I{row_idx}")
                        formula = formula.replace("H2", f"H{row_idx}")
                        formula = formula.replace("staff", f"{staff_}")
                        target_cell.value = formula
                    else:
                        target_cell.value = formula
                    if source_cell.has_style:
                        target_cell._style = source_cell._style
            else:
                print("hash_value not in existing_hash_keys")
       # sheet.cell(row=4, column=3).value = f"=AVERAGEIFS(C7:C19; I7:I19; I7)"  # Примерна формула за изчисляване на средна стойност
        cell = sheet.cell(row=4, column=3)
        cell.font = Font(color="000000")
        cell.value = "=AVERAGEIFS(C6:C{}, I6:I{}, I6)".format(sheet.max_row, sheet.max_row) # Примерна формула за изчисляване на средна стойност



        # # Създаване на графика
        # chart = LineChart()
        # chart.title = "Активна и Реактивна мощност по месеци"
        # chart.style = 13
        # chart.x_axis.title = "Месец"
        # chart.y_axis.title = "Мощност (W)"

        # # Данни за графиката
        # data = Reference(sheet, min_col=3, max_col=4, min_row=6, max_row=sheet.max_row)
        # categories = Reference(sheet, min_col=2, min_row=6, max_row=sheet.max_row)  # Колона B за "Месец"
        # chart.add_data(data, titles_from_data=True)
        # chart.set_categories(categories)

        # # Поставяне на графиката под таблицата
        # sheet.add_chart(chart, f"A{sheet.max_row + 2}")

    # Премахване на празния sheet, ако съществува
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Автоматично задаване на ширина на колоните според най-дългия текст
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value )) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                    if max_length > 20:  # Ограничаване на максималната ширина
                        max_length = 20
                except Exception:
                    pass
            sheet.column_dimensions[column].width = max_length + 2

    workbook.save("fill_"+excel_path)

def simulate_pdf_extraction(pdf_path):
    """Симулира извличане на текст от PDF (заменете с реална логика)."""
    return f"""
    Наименование на обекта: Обект 1
    Адрес на обекта: ул. Примерна 1
    Основание: Електрическа енергия за месец Януари 2025
    100 кВтч
    50 кВАрч
    """

def main():
    parser = argparse.ArgumentParser(description="Process multiple PDF files and save data to an Excel file.")
    parser.add_argument("excel_path", nargs="?", default="BookBase.xlsx", help="Path to the output Excel file.")
    parser.add_argument("pdf_directory", nargs="?", default="test", help="Path to the directory containing PDF files.")
    args = parser.parse_args()

    if not os.path.isdir(args.pdf_directory):
        print("The specified directory does not exist.")
        return

    # Обработка на PDF файловете
    data_by_object_code = process_pdfs(args.pdf_directory)

    # Генериране на Excel файла
    generate_excel(data_by_object_code, args.excel_path)
    # След като имаш data_by_object_code:
    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(data_by_object_code, f, ensure_ascii=False, default=str, indent=2)
if __name__ == "__main__":
    main()